import sys
import re
import os
import logging
import hashlib
import pyodbc
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QVBoxLayout,
    QPushButton, QWidget, QHBoxLayout, QAbstractItemView, QFileDialog, QMessageBox,
    QLineEdit, QLabel, QHeaderView, QAction, QFrame, QMenu, QDialog, QFormLayout,
    QComboBox, QGroupBox, QTabWidget, QTextEdit, QCheckBox, QListWidget, QListWidgetItem,
    QInputDialog, QDialogButtonBox, QSizePolicy
)
from PyQt5.QtCore import Qt, QSize
from PyQt5.QtGui import QIcon, QFont
import pandas as pd


def validate_course_number(cno):
    """验证课程号格式：一个字母加三个数字"""
    pattern = re.compile(r'^[A-Za-z]\d{3}$')
    if not pattern.match(cno):
        return False, "课程号必须为一个字母后跟三个数字（例如：C001）"
    return True, ""

def __init__(self):
    super().__init__()

    # 设置窗口基本属性
    self.setWindowTitle("学生信息管理系统")
    self.setGeometry(100, 100, 1200, 700)

    self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

    # 设置应用图标 - 使用相对路径
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "image", "pic.ico")
    if os.path.exists(icon_path):
        self.setWindowIcon(QIcon(icon_path))
    else:
        logging.warning(f"图标文件未找到: {icon_path}")


# 配置日志记录
os.makedirs('logs', exist_ok=True)
logging.basicConfig(
    filename='logs/student_app.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

# 数据库连接配置
DB_CONFIG = {
    'server': 'localhost',  # 服务器地址
    'database': 'student',  # 数据库名称
    'trusted_connection': 'no',  # 是否使用Windows身份验证
    'driver': '{ODBC Driver 17 for SQL Server}'  # ODBC驱动
}

# 全局常量
MAX_GRADE = 100  # 最高成绩
MIN_GRADE = 0  # 最低成绩


class SearchableComboBox(QComboBox):
    """可搜索的下拉框"""

    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.setMaxVisibleItems(15)  # 显示15个选项

        # 设置过滤器
        self.completer = self.completer()
        self.completer.setFilterMode(Qt.MatchContains)
        self.completer.setCaseSensitivity(Qt.CaseInsensitive)

        # 连接信号
        self.lineEdit().textEdited.connect(self.on_text_edited)

    def on_text_edited(self, text):
        """文本编辑时过滤项目"""
        if not text:
            self.showPopup()
            return

        self.completer.setCompletionPrefix(text)
        self.showPopup()


class AddScoreDialog(QDialog):
    """添加学生成绩对话框 """

    def __init__(self, parent=None, cursor=None):
        super().__init__(parent)
        self.cursor = cursor
        self.setWindowTitle("添加学生成绩")
        self.setMinimumWidth(500)
        self.setMinimumHeight(500)
        self.setModal(True)

        # 移除问号图标
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        if hasattr(parent, 'apply_dialog_style'):
            parent.apply_dialog_style(self)

        # 创建布局
        main_layout = QVBoxLayout(self)

        # 添加选项卡
        tabs = QTabWidget()
        student_tab = QWidget()
        course_tab = QWidget()

        # 设置学生选择界面
        student_layout = QVBoxLayout(student_tab)

        # 学生搜索框
        student_search_layout = QHBoxLayout()
        student_search_label = QLabel("搜索学生:")
        self.student_search = QLineEdit()
        self.student_search.setPlaceholderText("输入学号或姓名搜索...")

        student_search_layout.addWidget(student_search_label)
        student_search_layout.addWidget(self.student_search)

        # 学生列表
        self.student_list = QListWidget()
        self.student_list.setSelectionMode(QAbstractItemView.SingleSelection)

        # 添加到布局
        student_layout.addLayout(student_search_layout)
        student_layout.addWidget(QLabel("选择学生:"))
        student_layout.addWidget(self.student_list)

        # 设置课程选择界面
        course_layout = QVBoxLayout(course_tab)

        # 课程搜索框
        course_search_layout = QHBoxLayout()
        course_search_label = QLabel("搜索课程:")
        self.course_search = QLineEdit()
        self.course_search.setPlaceholderText("输入课程号或名称搜索...")

        course_search_layout.addWidget(course_search_label)
        course_search_layout.addWidget(self.course_search)

        # 课程列表
        self.course_list = QListWidget()
        self.course_list.setSelectionMode(QAbstractItemView.SingleSelection)

        # 添加到布局
        course_layout.addLayout(course_search_layout)
        course_layout.addWidget(QLabel("选择课程:"))
        course_layout.addWidget(self.course_list)

        # 添加选项卡
        tabs.addTab(student_tab, "选择学生")
        tabs.addTab(course_tab, "选择课程")

        # 成绩输入
        score_layout = QFormLayout()
        self.grade_input = QLineEdit()
        self.grade_input.setPlaceholderText("输入0-100之间的成绩")
        score_layout.addRow("成绩:", self.grade_input)

        # 选择信息显示
        self.selected_info = QLabel("请先在上方选择学生和课程")
        self.selected_info.setStyleSheet("color: #666; font-style: italic;")

        # 添加到主布局
        main_layout.addWidget(tabs)
        main_layout.addWidget(QLabel("当前选择:"))
        main_layout.addWidget(self.selected_info)
        main_layout.addLayout(score_layout)

        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        main_layout.addWidget(buttons)

        parent.setup_standard_buttons(buttons)

        # 加载数据
        self.load_students_data()
        self.load_courses_data()

        # 连接信号
        self.student_search.textChanged.connect(self.filter_students)
        self.course_search.textChanged.connect(self.filter_courses)
        self.student_list.itemSelectionChanged.connect(self.update_selected_info)
        self.course_list.itemSelectionChanged.connect(self.update_selected_info)

        # 存储数据
        self.selected_student = None
        self.selected_course = None

    def load_students_data(self):
        """加载学生数据到列表"""
        try:
            self.cursor.execute("SELECT Sno, name FROM Student ORDER BY Sno")
            students = self.cursor.fetchall()

            for student in students:
                sno = student[0].strip()
                name = student[1].strip()
                item = QListWidgetItem(f"{name} ({sno})")
                item.setData(Qt.UserRole, sno)
                self.student_list.addItem(item)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载学生数据失败: {e}")

    def load_courses_data(self):
        """加载课程数据到列表"""
        try:
            self.cursor.execute("SELECT Cno, course_name FROM Course ORDER BY Cno")
            courses = self.cursor.fetchall()

            for course in courses:
                cno = course[0].strip()
                name = course[1].strip()
                item = QListWidgetItem(f"{name} ({cno})")
                item.setData(Qt.UserRole, cno)
                self.course_list.addItem(item)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载课程数据失败: {e}")

    def filter_students(self, text):
        """过滤学生列表"""
        text = text.lower()
        for i in range(self.student_list.count()):
            item = self.student_list.item(i)
            if text in item.text().lower() or not text:
                item.setHidden(False)
            else:
                item.setHidden(True)

    def filter_courses(self, text):
        """过滤课程列表"""
        text = text.lower()
        for i in range(self.course_list.count()):
            item = self.course_list.item(i)
            if text in item.text().lower() or not text:
                item.setHidden(False)
            else:
                item.setHidden(True)

    def update_selected_info(self):
        """更新选择信息"""
        selected_student_items = self.student_list.selectedItems()
        selected_course_items = self.course_list.selectedItems()

        student_text = "未选择学生"
        course_text = "未选择课程"

        if selected_student_items:
            item = selected_student_items[0]
            student_text = item.text()
            self.selected_student = item.data(Qt.UserRole)

        if selected_course_items:
            item = selected_course_items[0]
            course_text = item.text()
            self.selected_course = item.data(Qt.UserRole)

        self.selected_info.setText(f"学生: {student_text} | 课程: {course_text}")

    def get_score_data(self):
        """获取成绩数据"""
        if not self.selected_student or not self.selected_course:
            QMessageBox.warning(self, "输入错误", "请先选择学生和课程")
            return None

        try:
            grade = float(self.grade_input.text())
            if grade < 0 or grade > 100:
                raise ValueError("成绩必须在0-100之间")

            return {
                'sno': self.selected_student,
                'cno': self.selected_course,
                'grade': grade
            }
        except ValueError as e:
            QMessageBox.warning(self, "输入错误", str(e))
            return None


class LoginDialog(QDialog):
    """登录对话框"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("登录")
        self.resize(350, 200)
        self.setModal(True)

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        # 创建布局
        layout = QVBoxLayout(self)

        # 服务器设置
        server_group = QGroupBox("数据库连接")
        server_layout = QFormLayout()

        self.server_edit = QLineEdit(DB_CONFIG['server'])
        self.database_edit = QLineEdit(DB_CONFIG['database'])

        server_layout.addRow("服务器:", self.server_edit)
        server_layout.addRow("数据库:", self.database_edit)
        server_group.setLayout(server_layout)

        # 用户验证
        auth_group = QGroupBox("用户验证")
        auth_layout = QFormLayout()

        self.username_edit = QLineEdit()
        self.password_edit = QLineEdit()
        self.password_edit.setEchoMode(QLineEdit.Password)

        auth_layout.addRow("用户名:", self.username_edit)
        auth_layout.addRow("密码:", self.password_edit)
        auth_group.setLayout(auth_layout)

        # 添加到主布局
        layout.addWidget(server_group)
        layout.addWidget(auth_group)

        # 按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        self.parent().setup_standard_buttons(button_box)  # 调用父窗口中的方法
        # 设置样式
        self.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            QPushButton {
                padding: 8px 16px;
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QGroupBox {
                font-weight: bold;
            }
        """)

        layout.addWidget(button_box)

    def get_credentials(self):
        """获取登录凭据"""
        return {
            'server': self.server_edit.text().strip(),
            'database': self.database_edit.text().strip(),
            'username': self.username_edit.text().strip(),
            'password': self.password_edit.text()
        }


class StudentManagementSystem(QMainWindow):
    def __init__(self):
        super().__init__()

        # 设置窗口基本属性
        self.setWindowTitle("学生信息管理系统")
        self.setGeometry(100, 100, 1200, 700)

        self.setWindowFlags(self.windowFlags() & ~Qt.WindowContextHelpButtonHint)

        # 设置全局字体
        font = QFont("Microsoft YaHei", 10)
        font.setStyleHint(QFont.SansSerif)
        QApplication.setFont(font)

        # 显示登录对话框
        if not self.show_login_dialog():
            # 登录失败，退出应用
            sys.exit(1)

        # 初始化UI
        self.init_ui()

        # 加载学生数据
        self.load_students()

    def apply_dialog_style(self, dialog):
        """为对话框和按钮应用统一样式，增强文字可见性"""
        # 先设置对话框基本样式
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #000000;
                font-weight: bold;
            }
            QLineEdit {
                color: #000000;
                background-color: white;
                border: 1px solid #aaa;
                padding: 5px;
            }
        """)

        # 找到并修改所有QPushButton，包括QDialogButtonBox中的按钮
        for button in dialog.findChildren(QPushButton):
            # 完全覆盖任何继承的样式
            button.setStyleSheet("""
                QPushButton {
                    color: #000000;
                    background-color: #e0e0e0;
                    font-weight: bold;
                    min-width: 80px;
                    padding: 5px 15px;
                    border: 2px solid #888;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #d0d0d0;
                    border: 2px solid #666;
                }
                QPushButton:pressed {
                    background-color: #c0c0c0;
                    padding-top: 6px;
                    padding-bottom: 4px;
                }
            """)

        # 特别处理OK和Cancel按钮，使其更加醒目
        for button in dialog.findChildren(QPushButton):
            text = button.text().lower()
            if "ok" in text or "确定" in text or "是" in text or "保存" in text:
                # 强调OK/确定按钮
                button.setStyleSheet("""
                    QPushButton {
                        color: #ffffff;
                        background-color: #4caf50;
                        font-weight: bold;
                        min-width: 80px;
                        padding: 5px 15px;
                        border: 2px solid #388e3c;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #388e3c;
                        border: 2px solid #2e7d32;
                    }
                    QPushButton:pressed {
                        background-color: #2e7d32;
                        padding-top: 6px;
                        padding-bottom: 4px;
                    }
                """)
            elif "cancel" in text or "取消" in text or "否" in text:
                # 设置取消按钮样式
                button.setStyleSheet("""
                    QPushButton {
                        color: #000000;
                        background-color: #e0e0e0;
                        font-weight: bold;
                        min-width: 80px;
                        padding: 5px 15px;
                        border: 2px solid #9e9e9e;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #d0d0d0;
                        border: 2px solid #757575;
                    }
                    QPushButton:pressed {
                        background-color: #c0c0c0;
                        padding-top: 6px;
                        padding-bottom: 4px;
                    }
                """)

    def show_login_dialog(self):
        """显示登录对话框并验证凭据"""
        dialog = LoginDialog(self)

        if dialog.exec_() == QDialog.Accepted:
            credentials = dialog.get_credentials()

            # 更新全局配置
            DB_CONFIG['server'] = credentials['server']
            DB_CONFIG['database'] = credentials['database']

            try:
                # 尝试连接数据库
                conn_str = f"DRIVER={DB_CONFIG['driver']};SERVER={DB_CONFIG['server']};DATABASE={DB_CONFIG['database']};UID={credentials['username']};PWD={credentials['password']}"
                self.conn = pyodbc.connect(conn_str)
                self.cursor = self.conn.cursor()

                # 记录登录
                self.current_user = credentials['username']
                logging.info(f"用户 {self.current_user} 登录成功")

                # 检查用户表是否存在，更新最后登录时间
                try:
                    self.cursor.execute("UPDATE Users SET last_login = GETDATE() WHERE username = ?",
                                        (self.current_user,))
                    self.conn.commit()
                except:
                    # 用户表可能不存在，忽略错误
                    pass

                return True

            except pyodbc.Error as e:
                logging.error(f"数据库连接错误: {e}")

                # 错误代码细分处理
                error_code = str(e).split('(')[1].split(')')[0] if '(' in str(e) and ')' in str(e) else "未知"
                error_msg = ""

                if "18456" in str(e):
                    error_msg = "用户名或密码错误，请检查登录凭据"
                elif "28000" in str(e):
                    error_msg = "SQL Server登录失败，请检查用户账号权限"
                elif "08001" in str(e):
                    error_msg = "无法连接到服务器，请检查服务器地址或网络连接"
                elif "42000" in str(e):
                    error_msg = "数据库访问被拒绝，请检查权限设置"
                elif "HYT00" in str(e):
                    error_msg = "连接超时，请检查网络或服务器是否运行"
                elif "01000" in str(e):
                    error_msg = "数据库不存在或无法访问，请检查数据库名称"
                else:
                    error_msg = f"连接失败: {str(e)}"

                QMessageBox.critical(self, "连接错误", f"{error_msg}\n\n错误代码: {error_code}")
                return False

        return False

    def init_ui(self):
        """初始化用户界面"""
        self.set_global_styles()

        # 创建中央部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 创建标题标签
        header_label = QLabel("学生信息管理系统", self)
        header_label.setStyleSheet("font-size: 24px; font-weight: bold; padding: 15px; color: #2e7d32;")
        header_label.setAlignment(Qt.AlignCenter)

        # 创建选项卡
        tab_widget = QTabWidget()

        # 学生成绩标签页
        student_score_tab = QWidget()
        student_score_layout = QVBoxLayout(student_score_tab)

        # 搜索栏
        self.create_search_bar(student_score_layout)

        # 学生表格
        self.create_student_table(student_score_layout)

        # 按钮区域
        self.create_buttons(student_score_layout)

        # 学生管理标签页
        student_tab = QWidget()
        student_layout = QVBoxLayout(student_tab)
        self.create_student_management_tab(student_layout)

        # 课程信息标签页
        course_tab = QWidget()
        course_layout = QVBoxLayout(course_tab)
        self.create_course_tab(course_layout)

        # 添加标签页
        tab_widget.addTab(student_score_tab, "学生成绩")
        tab_widget.addTab(student_tab, "学生管理")
        tab_widget.addTab(course_tab, "课程管理")

        # 添加组件到主布局
        main_layout.addWidget(header_label)
        main_layout.addWidget(tab_widget)

        # 创建状态栏
        self.statusBar().showMessage(
            f"已连接到 {DB_CONFIG['server']}/{DB_CONFIG['database']} - 用户: {self.current_user}")
        self.statusBar().setStyleSheet("font-size: 12px; padding: 4px;")

        # 初始化菜单
        self.init_menu()

    def set_global_styles(self):
        """设置全局样式表"""
        button_style = """
                QPushButton {
                    padding: 6px 12px;
                    border-radius: 4px;
                    /* 移除 color: white; 这一行 */
                    border: none;
                    box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
                }
                QPushButton:hover {
                    background-color: rgba(0, 0, 0, 0.1); 
                }
                QPushButton:pressed {
                    box-shadow: 0 1px 2px rgba(0, 0, 0, 0.2);
                    margin-top: 1px;
                    margin-bottom: -1px;
                }
                QPushButton:focus {
                    outline: none;
                    border: 1px solid white;
                }
            """
        self.setStyleSheet(button_style)

    def create_search_bar(self, parent_layout):
        """创建高级搜索筛选"""
        search_group = QGroupBox("高级搜索筛选")
        search_layout = QVBoxLayout()

        # 创建基本搜索行
        basic_search_layout = QHBoxLayout()

        # 搜索字段选择器
        self.search_field = QComboBox()
        self.search_field.addItem("全部字段")
        self.search_field.currentIndexChanged.connect(self.update_search_operator)

        # 搜索操作符选择器
        self.search_operator = QComboBox()
        self.search_operator.addItem("包含")

        # 搜索输入框
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入关键词搜索...")

        # 搜索按钮
        self.search_btn = self.create_button("搜索", "#2196f3")
        # 重置按钮
        self.reset_btn = self.create_button("重置", "#9e9e9e")

        # 连接信号
        self.search_btn.clicked.connect(self.search_students)
        self.reset_btn.clicked.connect(self.reset_search)
        self.search_input.returnPressed.connect(self.search_students)

        # 基本搜索行布局
        basic_search_layout.addWidget(QLabel("搜索字段:"))
        basic_search_layout.addWidget(self.search_field)
        basic_search_layout.addWidget(QLabel("条件:"))
        basic_search_layout.addWidget(self.search_operator)
        basic_search_layout.addWidget(QLabel("关键词:"))
        basic_search_layout.addWidget(self.search_input, 1)  # 搜索框伸展
        basic_search_layout.addWidget(self.search_btn)
        basic_search_layout.addWidget(self.reset_btn)

        # 添加高级筛选选项
        advanced_layout = QHBoxLayout()
        self.advanced_check = QCheckBox("启用高级筛选")
        self.advanced_check.stateChanged.connect(self.toggle_advanced_search)
        advanced_layout.addWidget(self.advanced_check)
        advanced_layout.addStretch()

        # 高级筛选区域 - 初始隐藏
        self.advanced_search_widget = QWidget()
        advanced_search_layout = QVBoxLayout(self.advanced_search_widget)

        # 课程筛选
        course_filter_layout = QHBoxLayout()
        course_filter_layout.addWidget(QLabel("选择课程:"))

        # 修改为普通下拉框，避免输入问题
        self.course_filter = QComboBox()
        self.course_filter.setMaxVisibleItems(15)  # 显示15个选项
        self.course_filter.addItem("所有课程")
        course_filter_layout.addWidget(self.course_filter, 1)

        # 成绩范围筛选
        grade_filter_layout = QHBoxLayout()
        grade_filter_layout.addWidget(QLabel("成绩范围:"))
        self.min_grade = QLineEdit()
        self.min_grade.setPlaceholderText("最低分")
        self.min_grade.setMaximumWidth(80)
        grade_filter_layout.addWidget(self.min_grade)
        grade_filter_layout.addWidget(QLabel("到"))
        self.max_grade = QLineEdit()
        self.max_grade.setPlaceholderText("最高分")
        self.max_grade.setMaximumWidth(80)
        grade_filter_layout.addWidget(self.max_grade)
        grade_filter_layout.addStretch()

        # 移除预设成绩范围按钮，改为简洁的下拉菜单
        grade_preset_layout = QHBoxLayout()
        grade_preset_layout.addWidget(QLabel("成绩等级:"))
        self.grade_preset = QComboBox()
        self.grade_preset.addItem("不筛选")
        self.grade_preset.addItems(["优秀(90-100)", "良好(80-89)", "中等(70-79)", "及格(60-69)", "不及格(<60)"])
        self.grade_preset.currentIndexChanged.connect(self.apply_grade_preset)
        grade_preset_layout.addWidget(self.grade_preset)
        grade_preset_layout.addStretch()

        # 排序选项
        sort_layout = QHBoxLayout()
        sort_layout.addWidget(QLabel("排序:"))
        self.sort_field = QComboBox()
        self.sort_field.addItems(["学号", "姓名", "课程号", "学分", "成绩"])
        sort_layout.addWidget(self.sort_field)

        self.sort_order = QComboBox()
        self.sort_order.addItems(["升序", "降序"])
        sort_layout.addWidget(self.sort_order)
        sort_layout.addStretch()

        # 应用按钮
        apply_layout = QHBoxLayout()
        self.apply_btn = self.create_button("应用筛选", "#4caf50")
        self.apply_btn.clicked.connect(self.apply_advanced_search)
        apply_layout.addStretch()
        apply_layout.addWidget(self.apply_btn)

        # 添加所有高级筛选组件
        advanced_search_layout.addLayout(course_filter_layout)
        advanced_search_layout.addLayout(grade_filter_layout)
        advanced_search_layout.addLayout(grade_preset_layout)
        advanced_search_layout.addLayout(sort_layout)
        advanced_search_layout.addLayout(apply_layout)

        # 初始隐藏高级筛选
        self.advanced_search_widget.setVisible(False)

        # 添加到主布局
        search_layout.addLayout(basic_search_layout)
        search_layout.addLayout(advanced_layout)
        search_layout.addWidget(self.advanced_search_widget)

        search_group.setLayout(search_layout)
        parent_layout.addWidget(search_group)

        # 加载课程数据到筛选下拉框
        self.load_course_filter_data()

    def update_search_operator(self):
        """根据选择的搜索字段更新操作符选项"""
        self.search_operator.clear()
        field_text = self.search_field.currentText()

        if field_text in ["学分", "成绩"]:
            # 数值字段
            self.search_operator.addItems(["等于", "大于", "小于", "大于等于", "小于等于", "不等于"])
        else:
            # 文本字段
            self.search_operator.addItems(["包含", "等于", "开头是", "结尾是"])

    def toggle_advanced_search(self, state):
        """切换高级搜索面板显示状态"""
        self.advanced_search_widget.setVisible(state == Qt.Checked)

        # 如果关闭高级搜索，重置高级搜索条件
        if state != Qt.Checked:
            self.course_filter.setCurrentIndex(0)
            self.min_grade.clear()
            self.max_grade.clear()
            self.grade_preset.setCurrentIndex(0)
            self.sort_field.setCurrentIndex(0)
            self.sort_order.setCurrentIndex(0)

    def load_course_filter_data(self):
        """加载课程数据到筛选下拉框"""
        try:
            self.cursor.execute("SELECT Cno, course_name FROM Course ORDER BY Cno")
            courses = self.cursor.fetchall()

            for course in courses:
                cno = course[0].strip()
                name = course[1].strip()
                self.course_filter.addItem(f"{name} ({cno})")
        except Exception as e:
            logging.error(f"加载课程筛选数据错误: {e}")

    def apply_grade_preset(self, index):
        """应用成绩预设"""
        if index == 0:  # "不筛选"
            self.min_grade.clear()
            self.max_grade.clear()
            return

        preset_text = self.grade_preset.currentText()

        if "90-100" in preset_text:
            self.min_grade.setText("90")
            self.max_grade.setText("100")
        elif "80-89" in preset_text:
            self.min_grade.setText("80")
            self.max_grade.setText("89")
        elif "70-79" in preset_text:
            self.min_grade.setText("70")
            self.max_grade.setText("79")
        elif "60-69" in preset_text:
            self.min_grade.setText("60")
            self.max_grade.setText("69")
        elif "<60" in preset_text:
            self.min_grade.setText("0")
            self.max_grade.setText("59")

    def apply_advanced_search(self):
        """应用高级搜索筛选"""
        # 确保高级搜索被激活
        if not self.advanced_check.isChecked():
            self.advanced_check.setChecked(True)

        # 执行搜索
        self.search_students()



    def search_students(self):
        """根据高级搜索条件筛选学生"""
        try:
            # 断开信号，防止加载数据时触发编辑信号
            self.table.itemChanged.disconnect(self.on_cell_changed)
        except:
            pass

        try:
            # 基本搜索参数
            search_text = self.search_input.text().strip()
            field_index = self.search_field.currentIndex()
            field_text = self.search_field.currentText()
            operator = self.search_operator.currentText()

            # 构建基本查询
            query = """
            SELECT s.Sno, s.name, sc.Cno, c.Credit, sc.Grade, c.course_name
            FROM Student_Score sc
            JOIN Student s ON sc.Sno = s.Sno
            JOIN Course c ON sc.Cno = c.Cno
            WHERE 1=1
            """

            params = []

            # 添加基本搜索条件
            if search_text:
                if field_index == 0:  # 全部字段
                    query += " AND (s.Sno LIKE ? OR s.name LIKE ? OR sc.Cno LIKE ? OR c.course_name LIKE ?)"
                    params.extend([f"%{search_text}%", f"%{search_text}%", f"%{search_text}%", f"%{search_text}%"])
                else:
                    # 根据字段类型和操作符构建条件
                    if field_text == "学号":
                        field_name = "s.Sno"
                    elif field_text == "姓名":
                        field_name = "s.name"
                    elif field_text == "课程号":
                        field_name = "sc.Cno"
                    elif field_text == "学分":
                        field_name = "c.Credit"
                    elif field_text == "成绩":
                        field_name = "sc.Grade"

                    # 构建条件
                    if field_text in ["学分", "成绩"]:
                        # 数值比较
                        try:
                            value = float(search_text)
                            if operator == "等于":
                                query += f" AND {field_name} = ?"
                            elif operator == "大于":
                                query += f" AND {field_name} > ?"
                            elif operator == "小于":
                                query += f" AND {field_name} < ?"
                            elif operator == "大于等于":
                                query += f" AND {field_name} >= ?"
                            elif operator == "小于等于":
                                query += f" AND {field_name} <= ?"
                            elif operator == "不等于":
                                query += f" AND {field_name} <> ?"
                            params.append(value)
                        except ValueError:
                            QMessageBox.warning(self, "输入错误", f"{field_text}必须是数字")
                            return
                    else:
                        # 文本比较
                        if operator == "包含":
                            query += f" AND {field_name} LIKE ?"
                            params.append(f"%{search_text}%")
                        elif operator == "等于":
                            query += f" AND {field_name} = ?"
                            params.append(search_text)
                        elif operator == "开头是":
                            query += f" AND {field_name} LIKE ?"
                            params.append(f"{search_text}%")
                        elif operator == "结尾是":
                            query += f" AND {field_name} LIKE ?"
                            params.append(f"%{search_text}")

            # 添加高级筛选条件，如果高级搜索面板可见
            if self.advanced_search_widget.isVisible():
                # 课程筛选
                selected_course = self.course_filter.currentText()
                if selected_course != "所有课程":
                    course_cno = selected_course.split("(")[-1].split(")")[0].strip()
                    query += " AND sc.Cno = ?"
                    params.append(course_cno)

                # 成绩范围筛选
                min_grade_text = self.min_grade.text().strip()
                max_grade_text = self.max_grade.text().strip()

                if min_grade_text:
                    try:
                        min_grade = float(min_grade_text)
                        query += " AND sc.Grade >= ?"
                        params.append(min_grade)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", "最低分必须是数字")
                        return

                if max_grade_text:
                    try:
                        max_grade = float(max_grade_text)
                        query += " AND sc.Grade <= ?"
                        params.append(max_grade)
                    except ValueError:
                        QMessageBox.warning(self, "输入错误", "最高分必须是数字")
                        return

                # 排序
                sort_field = self.sort_field.currentText()
                sort_order = "ASC" if self.sort_order.currentText() == "升序" else "DESC"

                if sort_field == "学号":
                    query += f" ORDER BY s.Sno {sort_order}"
                elif sort_field == "姓名":
                    query += f" ORDER BY s.name {sort_order}"
                elif sort_field == "课程号":
                    query += f" ORDER BY sc.Cno {sort_order}"
                elif sort_field == "学分":
                    query += f" ORDER BY c.Credit {sort_order}"
                elif sort_field == "成绩":
                    query += f" ORDER BY sc.Grade {sort_order}"
            else:
                # 默认排序
                query += " ORDER BY s.Sno, sc.Cno"

            # 执行查询
            self.cursor.execute(query, params)
            records = self.cursor.fetchall()

            # 清空表格
            self.table.setRowCount(0)

            # 添加数据到表格
            for i, record in enumerate(records):
                self.table.insertRow(i)
                # 只显示前5列（不显示course_name）
                for j in range(5):
                    value = record[j]
                    # 修复空格问题
                    if j in [0, 1, 2] and value is not None:
                        value = str(value).strip()
                    self.table.setItem(i, j, QTableWidgetItem(str(value)))

            # 更新状态栏
            self.statusBar().showMessage(f"找到 {len(records)} 条匹配记录")

        except pyodbc.Error as e:
            logging.error(f"搜索学生数据错误: {e}")
            QMessageBox.warning(self, "错误", f"搜索学生数据失败: {e}")

        # 重新连接信号
        self.table.itemChanged.connect(self.on_cell_changed)

    def create_student_table(self, parent_layout):
        """创建学生表格"""
        self.table = QTableWidget(0, 5)  # 5列：学号、姓名、课程号、学分、成绩
        self.table.setHorizontalHeaderLabels(["学号", "姓名", "课程号", "学分", "成绩"])

        # 设置表格属性
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        # 设置列宽
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 连接单元格编辑信号
        self.table.itemChanged.connect(self.on_cell_changed)

        parent_layout.addWidget(self.table)

    def create_student_management_tab(self, parent_layout):
        """创建学生管理标签页"""
        # 学生表格
        self.student_table = QTableWidget(0, 3)  # 3列：学号、姓名、操作按钮
        self.student_table.setHorizontalHeaderLabels(["学号", "姓名", "已修课程数"])

        # 设置表格属性
        self.student_table.setEditTriggers(QAbstractItemView.NoEditTriggers)  # 禁止直接编辑
        self.student_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.student_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.student_table.setAlternatingRowColors(True)
        self.student_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.student_table.customContextMenuRequested.connect(self.show_student_context_menu)

        # 设置列宽
        self.student_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 搜索区域
        search_layout = QHBoxLayout()
        self.student_search_input = QLineEdit()
        self.student_search_input.setPlaceholderText("输入学号或姓名搜索...")
        self.student_search_input.returnPressed.connect(self.search_student)

        search_btn = self.create_button("搜索", "#2196f3")
        search_btn.clicked.connect(self.search_student)

        reset_btn = self.create_button("重置", "#9e9e9e")
        reset_btn.clicked.connect(self.reset_student_search)

        search_layout.addWidget(QLabel("搜索:"))
        search_layout.addWidget(self.student_search_input, 1)
        search_layout.addWidget(search_btn)
        search_layout.addWidget(reset_btn)

        # 按钮区域 - 优化布局
        button_group = QGroupBox("操作")
        button_layout = QHBoxLayout()

        # 创建按钮并分组
        add_student_btn = self.create_button("添加学生", "#4caf50")
        edit_student_btn = self.create_button("编辑学生", "#2196f3")
        delete_student_btn = self.create_button("删除学生", "#f44336")
        view_score_btn = self.create_button("查看成绩", "#ff9800")
        refresh_student_btn = self.create_button("刷新", "#9c27b0")

        # 设置按钮大小策略
        for btn in [add_student_btn, edit_student_btn, delete_student_btn, view_score_btn, refresh_student_btn]:
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # 添加到布局
        button_layout.addWidget(add_student_btn)
        button_layout.addWidget(edit_student_btn)
        button_layout.addWidget(view_score_btn)
        button_layout.addWidget(delete_student_btn)
        button_layout.addStretch(1)
        button_layout.addWidget(refresh_student_btn)

        button_group.setLayout(button_layout)

        # 连接信号
        add_student_btn.clicked.connect(self.add_student)
        edit_student_btn.clicked.connect(self.edit_student)
        delete_student_btn.clicked.connect(self.delete_student)
        view_score_btn.clicked.connect(self.view_student_scores)
        refresh_student_btn.clicked.connect(self.load_student_list)

        # 添加到主布局
        parent_layout.addLayout(search_layout)
        parent_layout.addWidget(self.student_table)
        parent_layout.addWidget(button_group)

        # 加载学生数据
        self.load_student_list()

    def create_course_tab(self, parent_layout):
        """创建课程管理标签页"""
        # 课程表格
        self.course_table = QTableWidget(0, 3)  # 3列：课程号、课程名称、学分
        self.course_table.setHorizontalHeaderLabels(["课程号", "课程名称", "学分"])

        # 设置表格属性
        self.course_table.setEditTriggers(QAbstractItemView.DoubleClicked)
        self.course_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.course_table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.course_table.setAlternatingRowColors(True)
        # 添加右键菜单支持
        self.course_table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.course_table.customContextMenuRequested.connect(self.show_course_context_menu)

        # 设置列宽
        self.course_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 搜索区域
        search_layout = QHBoxLayout()
        self.course_search_input = QLineEdit()
        self.course_search_input.setPlaceholderText("输入课程号或课程名称搜索...")
        self.course_search_input.returnPressed.connect(self.search_course)

        search_btn = self.create_button("搜索", "#2196f3")
        search_btn.clicked.connect(self.search_course)

        reset_btn = self.create_button("重置", "#9e9e9e")
        reset_btn.clicked.connect(self.reset_course_search)

        search_layout.addWidget(QLabel("搜索:"))
        search_layout.addWidget(self.course_search_input, 1)
        search_layout.addWidget(search_btn)
        search_layout.addWidget(reset_btn)

        # 按钮区域 - 优化布局
        button_group = QGroupBox("操作")
        button_layout = QHBoxLayout()

        # 创建按钮并分组
        add_course_btn = self.create_button("添加课程", "#4caf50")
        edit_course_btn = self.create_button("编辑课程", "#2196f3")
        delete_course_btn = self.create_button("删除课程", "#f44336")
        refresh_course_btn = self.create_button("刷新", "#9c27b0")

        # 设置按钮大小策略
        for btn in [add_course_btn, edit_course_btn, delete_course_btn, refresh_course_btn]:
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        # 添加到布局
        button_layout.addWidget(add_course_btn)
        button_layout.addWidget(edit_course_btn)
        button_layout.addWidget(delete_course_btn)
        button_layout.addStretch(1)
        button_layout.addWidget(refresh_course_btn)

        button_group.setLayout(button_layout)

        # 连接信号
        add_course_btn.clicked.connect(self.add_course)
        edit_course_btn.clicked.connect(self.edit_course)
        delete_course_btn.clicked.connect(self.delete_course)
        refresh_course_btn.clicked.connect(self.load_courses)

        # 添加到主布局
        parent_layout.addLayout(search_layout)
        parent_layout.addWidget(self.course_table)
        parent_layout.addWidget(button_group)

        # 加载课程数据
        self.load_courses()

    def setup_standard_buttons(self, buttonBox):
        """设置标准按钮样式，确保可见性"""
        for button in buttonBox.buttons():
            role = buttonBox.buttonRole(button)

            # 针对接受按钮(OK, Yes, Apply等)
            if role == QDialogButtonBox.AcceptRole or role == QDialogButtonBox.YesRole:
                button.setStyleSheet("""
                    QPushButton {
                        color: #ffffff;
                        background-color: #4caf50;
                        font-weight: bold;
                        min-width: 80px;
                        padding: 5px 15px;
                        border: 2px solid #388e3c;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #388e3c;
                    }
                    QPushButton:pressed {
                        padding-top: 6px;
                        padding-bottom: 4px;
                    }
                """)
            # 针对拒绝按钮(Cancel, No等)
            elif role == QDialogButtonBox.RejectRole or role == QDialogButtonBox.NoRole:
                button.setStyleSheet("""
                    QPushButton {
                        color: #000000;
                        background-color: #e0e0e0;
                        font-weight: bold;
                        min-width: 80px;
                        padding: 5px 15px;
                        border: 2px solid #9e9e9e;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #d0d0d0;
                    }
                    QPushButton:pressed {
                        padding-top: 6px;
                        padding-bottom: 4px;
                    }
                """)
            else:
                # 其他按钮
                button.setStyleSheet("""
                    QPushButton {
                        color: #000000;
                        background-color: #e0e0e0;
                        font-weight: bold;
                        min-width: 80px;
                        padding: 5px 15px;
                        border: 2px solid #9e9e9e;
                        border-radius: 4px;
                    }
                    QPushButton:hover {
                        background-color: #d0d0d0;
                    }
                    QPushButton:pressed {
                        padding-top: 6px;
                        padding-bottom: 4px;
                    }
                """)

    def create_buttons(self, parent_layout):
        """创建按钮区域 - 优化布局"""
        # 创建按钮组
        button_group = QGroupBox("操作")
        btn_layout = QHBoxLayout()

        # 基本操作按钮
        operations_layout = QHBoxLayout()
        self.add_btn = self.create_button("添加成绩", "#4caf50")
        self.edit_btn = self.create_button("编辑成绩", "#2196f3")
        self.delete_btn = self.create_button("删除成绩", "#f44336")
        self.save_btn = self.create_button("保存修改", "#9c27b0")
        self.refresh_btn = self.create_button("刷新数据", "#607d8b")

        # 设置按钮大小策略
        for btn in [self.add_btn, self.edit_btn, self.delete_btn, self.save_btn, self.refresh_btn]:
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        operations_layout.addWidget(self.add_btn)
        operations_layout.addWidget(self.edit_btn)
        operations_layout.addWidget(self.delete_btn)
        operations_layout.addWidget(self.save_btn)
        operations_layout.addWidget(self.refresh_btn)

        # 导入导出按钮
        import_export_layout = QHBoxLayout()
        self.import_btn = self.create_button("导入Excel", "#ff9800")
        self.export_btn = self.create_button("导出Excel", "#795548")

        self.import_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.export_btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        import_export_layout.addWidget(self.import_btn)
        import_export_layout.addWidget(self.export_btn)

        # 添加到主布局
        btn_layout.addLayout(operations_layout, 3)  # 操作按钮占3份
        btn_layout.addStretch(1)  # 弹性空间
        btn_layout.addLayout(import_export_layout, 1)  # 导入导出按钮占1份

        button_group.setLayout(btn_layout)
        parent_layout.addWidget(button_group)

        # 连接信号槽
        self.add_btn.clicked.connect(self.add_student_score)
        self.edit_btn.clicked.connect(self.edit_student_score)
        self.delete_btn.clicked.connect(self.delete_student_score)
        self.save_btn.clicked.connect(self.save_all_changes)
        self.refresh_btn.clicked.connect(self.load_students)
        self.import_btn.clicked.connect(self.import_excel)
        self.export_btn.clicked.connect(self.export_excel)

    def init_menu(self):
        """初始化菜单栏 - 删除管理菜单"""
        menubar = self.menuBar()

        # 文件菜单
        file_menu = menubar.addMenu("文件")

        import_action = QAction("导入Excel", self)
        import_action.triggered.connect(self.import_excel)
        file_menu.addAction(import_action)

        export_action = QAction("导出Excel", self)
        export_action.triggered.connect(self.export_excel)
        file_menu.addAction(export_action)

        file_menu.addSeparator()

        exit_action = QAction("退出", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # 帮助菜单
        help_menu = menubar.addMenu("帮助")

        about_action = QAction("关于", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)

    def create_button(self, text, color):
        """创建美观的按钮"""
        btn = QPushButton(text)

        # 根据背景颜色自动选择文字颜色
        # 对于深色背景使用白色文字，浅色背景使用黑色文字
        text_color = "white"
        if color in ["#e0e0e0", "#f5f5f5", "#ffffff", "#f0f0f0", "#eeeeee", "#dddddd"]:
            text_color = "black"

        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color}; 
                color: {text_color}; 
                padding: 8px; 
                font-size: 13px;
                font-weight: bold;
                border-radius: 4px;
                min-width: 80px;
                border: 2px solid rgba(0, 0, 0, 0.2);
            }}
            QPushButton:hover {{
                background-color: {color};
                border: 2px solid rgba(0, 0, 0, 0.4);
            }}
            QPushButton:pressed {{
                background-color: {color};  
                padding-top: 9px;
                padding-bottom: 7px;
            }}
            QPushButton:focus {{
                outline: none;
                border: 2px solid rgba(0, 0, 0, 0.5);
            }}
            QPushButton:disabled {{
                background-color: #cccccc;
                color: #666666;
                border: 1px solid #999999;
            }}
        """)
        return btn

    def load_students(self):
        """从数据库加载学生成绩数据"""
        try:
            # 断开信号，防止加载数据时触发编辑信号
            self.table.itemChanged.disconnect(self.on_cell_changed)
        except:
            pass

        try:
            # 执行查询，获取学生成绩数据
            query = """
            SELECT s.Sno, s.name, sc.Cno, c.Credit, sc.Grade
            FROM Student_Score sc
            JOIN Student s ON sc.Sno = s.Sno
            JOIN Course c ON sc.Cno = c.Cno
            ORDER BY s.Sno, sc.Cno
            """
            self.cursor.execute(query)
            records = self.cursor.fetchall()

            # 清空表格
            self.table.setRowCount(0)

            # 添加数据到表格
            for i, record in enumerate(records):
                self.table.insertRow(i)
                for j, value in enumerate(record):
                    # 修复空格问题 - 字符串类型数据去除空格
                    if j in [0, 1, 2] and value is not None:  # 学号、姓名、课程号字段
                        value = str(value).strip()
                    self.table.setItem(i, j, QTableWidgetItem(str(value)))

            # 动态更新搜索下拉框
            # 首先保留"全部字段"选项，清除其他选项
            while self.search_field.count() > 1:
                self.search_field.removeItem(1)

            # 添加表格的列名到搜索下拉框
            for i in range(self.table.columnCount()):
                column_name = self.table.horizontalHeaderItem(i).text()
                self.search_field.addItem(column_name)

            # 更新状态栏
            self.statusBar().showMessage(f"已加载 {len(records)} 条学生成绩记录")

        except pyodbc.Error as e:
            logging.error(f"加载学生数据错误: {e}")
            QMessageBox.warning(self, "错误", f"加载学生数据失败: {e}")

        # 重新连接信号
        self.table.itemChanged.connect(self.on_cell_changed)

    def load_courses(self):
        """加载课程数据"""
        try:
            # 执行查询
            query = "SELECT Cno, course_name, Credit FROM Course ORDER BY Cno"
            self.cursor.execute(query)
            courses = self.cursor.fetchall()

            # 清空表格
            self.course_table.setRowCount(0)

            # 添加数据到表格
            for i, course in enumerate(courses):
                self.course_table.insertRow(i)
                for j, value in enumerate(course):
                    # 修复空格问题 - 字符串类型数据去除空格
                    if j in [0, 1] and value is not None:  # 课程号、课程名称字段
                        value = str(value).strip()
                    self.course_table.setItem(i, j, QTableWidgetItem(str(value)))

            # 更新状态栏
            self.statusBar().showMessage(f"已加载 {len(courses)} 门课程")

        except pyodbc.Error as e:
            logging.error(f"加载课程数据错误: {e}")
            QMessageBox.warning(self, "错误", f"加载课程数据失败: {e}")


    def reset_search(self):
        """重置所有搜索条件并加载所有学生"""
        # 重置基本搜索
        self.search_field.setCurrentIndex(0)
        self.search_input.clear()

        # 重置高级搜索
        self.advanced_check.setChecked(False)
        self.course_filter.setCurrentIndex(0)
        self.min_grade.clear()
        self.max_grade.clear()
        self.grade_preset.setCurrentIndex(0)
        self.sort_field.setCurrentIndex(0)
        self.sort_order.setCurrentIndex(0)

        # 加载所有学生
        self.load_students()

    def show_context_menu(self, position):
        """显示右键菜单"""
        # 只有选中行时才显示菜单
        if not self.table.selectedItems():
            return

        context_menu = QMenu(self)

        edit_action = context_menu.addAction("编辑成绩")
        delete_action = context_menu.addAction("删除成绩")

        # 连接信号
        edit_action.triggered.connect(self.edit_student_score)
        delete_action.triggered.connect(self.delete_student_score)

        # 显示菜单
        context_menu.exec_(self.table.mapToGlobal(position))

    def show_course_context_menu(self, position):
        """显示课程表格右键菜单"""
        # 只有选中行时才显示菜单
        if not self.course_table.selectedItems():
            return

        context_menu = QMenu(self)

        edit_action = context_menu.addAction("编辑课程")
        delete_action = context_menu.addAction("删除课程")
        view_students_action = context_menu.addAction("查看选课学生")

        # 连接信号
        edit_action.triggered.connect(self.edit_course)
        delete_action.triggered.connect(self.delete_course)
        view_students_action.triggered.connect(self.view_course_students)

        # 显示菜单
        context_menu.exec_(self.course_table.mapToGlobal(position))

    def view_course_students(self):
        """查看选择了该课程的学生"""
        selected_row = self.course_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一门课程")
            return

        cno = self.course_table.item(selected_row, 0).text().strip()
        course_name = self.course_table.item(selected_row, 1).text().strip()

        try:
            # 查询选择了该课程的学生
            query = """
            SELECT s.Sno, s.name, sc.Grade
            FROM Student_Score sc
            JOIN Student s ON sc.Sno = s.Sno
            WHERE sc.Cno = ?
            ORDER BY sc.Grade DESC, s.Sno
            """
            self.cursor.execute(query, (cno,))
            students = self.cursor.fetchall()

            if not students:
                QMessageBox.information(self, "提示", f"暂无学生选择课程 {course_name} ({cno})")
                return

            # 创建对话框
            dialog = QDialog(self)
            dialog.setWindowTitle(f"选择课程 {course_name} ({cno}) 的学生")
            dialog.setMinimumSize(500, 400)

            self.apply_dialog_style(dialog)
            layout = QVBoxLayout(dialog)

            # 学生表格
            table = QTableWidget(len(students), 3)
            table.setHorizontalHeaderLabels(["学号", "姓名", "成绩"])
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            # 添加数据
            highest_grade = 0
            lowest_grade = 100
            total_grade = 0

            for i, student in enumerate(students):
                for j, value in enumerate(student):
                    if value is not None:
                        value = str(value).strip()
                    table.setItem(i, j, QTableWidgetItem(str(value)))

                # 统计成绩
                grade = float(student[2])
                total_grade += grade
                highest_grade = max(highest_grade, grade)
                lowest_grade = min(lowest_grade, grade)

            layout.addWidget(table)

            # 统计信息
            stats_layout = QFormLayout()

            stats_layout.addRow("选课人数:", QLabel(str(len(students))))
            stats_layout.addRow("最高成绩:", QLabel(f"{highest_grade:.1f}"))
            stats_layout.addRow("最低成绩:", QLabel(f"{lowest_grade:.1f}"))
            stats_layout.addRow("平均成绩:", QLabel(f"{total_grade / len(students):.2f}"))

            layout.addLayout(stats_layout)

            # 关闭按钮
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)

            dialog.exec_()

        except pyodbc.Error as e:
            logging.error(f"查询课程学生错误: {e}")
            QMessageBox.critical(self, "错误", f"查询课程学生失败: {e}")

    def on_cell_changed(self, item):
        """单元格内容变化处理"""
        # 仅处理成绩列
        if item.column() == 4:  # 成绩列
            row = item.row()

            try:
                # 尝试将输入转换为浮点数
                value = float(item.text())

                # 验证成绩范围
                if value < 0 or value > 100:
                    raise ValueError("成绩必须在0-100之间")

                # 高亮修改的单元格
                item.setBackground(Qt.yellow)

                # 更新状态栏
                self.statusBar().showMessage("单元格已修改，请点击保存按钮应用更改")

                # 确保保存按钮可见且突出
                self.save_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #9c27b0; 
                        color: white; 
                        padding: 8px; 
                        font-size: 13px;
                        border-radius: 4px;
                        min-width: 80px;
                        font-weight: bold;
                        border: 2px solid #4a148c;
                    }
                    QPushButton:hover {
                        background-color: #7b1fa2;
                    }
                """)
            except ValueError as e:
                # 恢复原值
                QMessageBox.warning(self, "输入错误", str(e))

                # 断开信号防止递归
                try:
                    self.table.itemChanged.disconnect(self.on_cell_changed)
                except:
                    pass

                # 重新设置原值（仅当有原始数据时）
                try:
                    sno = self.table.item(row, 0).text().strip()
                    cno = self.table.item(row, 2).text().strip()
                    self.cursor.execute("""
                        SELECT Grade FROM Student_Score 
                        WHERE Sno = ? AND Cno = ?
                    """, (sno, cno))
                    original_grade = self.cursor.fetchone()[0]
                    self.table.setItem(row, 4, QTableWidgetItem(str(original_grade)))
                except:
                    # 如果无法获取原始值，则重新加载所有数据
                    self.load_students()

                # 重新连接信号
                self.table.itemChanged.connect(self.on_cell_changed)

    def save_all_changes(self):
        """保存所有修改"""
        try:
            # 记录修改数量
            changes_count = 0

            # 遍历表格中的所有行
            for row in range(self.table.rowCount()):
                # 获取单元格背景色，检查是否修改过
                grade_item = self.table.item(row, 4)
                if grade_item and grade_item.background().color() == Qt.yellow:
                    try:
                        # 获取行数据
                        sno = self.table.item(row, 0).text().strip()
                        cno = self.table.item(row, 2).text().strip()
                        grade = float(grade_item.text())

                        # 更新数据库
                        query = """
                        UPDATE Student_Score
                        SET Grade = ?
                        WHERE Sno = ? AND Cno = ?
                        """
                        self.cursor.execute(query, (grade, sno, cno))
                        changes_count += 1
                    except (ValueError, TypeError) as e:
                        # 处理数据转换错误
                        QMessageBox.warning(self, "数据错误", f"第 {row + 1} 行成绩格式错误: {e}")
                        continue

            if changes_count > 0:
                # 提交事务
                self.conn.commit()

                # 重新加载数据（清除黄色标记）
                self.load_students()

                # 更新状态栏
                QMessageBox.information(self, "成功", f"已保存 {changes_count} 处修改")
                self.statusBar().showMessage(f"已保存 {changes_count} 处修改")
            else:
                QMessageBox.information(self, "提示", "没有检测到修改项")
                self.statusBar().showMessage("没有检测到修改项")

        except pyodbc.Error as e:
            logging.error(f"保存修改错误: {e}")
            QMessageBox.critical(self, "保存错误", f"保存修改失败: {e}")
            # 尝试回滚事务
            try:
                self.conn.rollback()
            except:
                pass

    def add_student_score(self):
        """添加学生成绩"""
        dialog = AddScoreDialog(self, self.cursor)

        self.apply_dialog_style(dialog)
        if dialog.exec_() == QDialog.Accepted:
            score_data = dialog.get_score_data()

            if score_data:
                try:
                    # 检查是否已存在该学生的该课程成绩
                    self.cursor.execute("""
                        SELECT COUNT(*) FROM Student_Score
                        WHERE Sno = ? AND Cno = ?
                    """, (score_data['sno'], score_data['cno']))

                    count = self.cursor.fetchone()[0]

                    if count > 0:
                        QMessageBox.warning(self, "添加失败", "该学生已有该课程的成绩记录")
                        return

                    # 添加新成绩
                    self.cursor.execute("""
                        INSERT INTO Student_Score (Sno, Cno, Grade)
                        VALUES (?, ?, ?)
                    """, (score_data['sno'], score_data['cno'], score_data['grade']))

                    self.conn.commit()

                    # 刷新数据
                    self.load_students()

                    # 更新状态栏
                    self.statusBar().showMessage("已添加新成绩记录")

                except pyodbc.Error as e:
                    self.conn.rollback()
                    logging.error(f"添加成绩错误: {e}")
                    QMessageBox.critical(self, "错误", f"添加成绩失败: {e}")

    def edit_student_score(self):
        """编辑学生成绩"""
        # 获取选中行
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一条记录")
            return

        # 获取当前成绩
        sno = self.table.item(selected_row, 0).text().strip()
        name = self.table.item(selected_row, 1).text().strip()
        cno = self.table.item(selected_row, 2).text().strip()
        current_grade = self.table.item(selected_row, 4).text()

        # 显示对话框获取新成绩
        new_grade, ok = QInputDialog.getDouble(
            self, f"编辑成绩 - {name}",
            f"请输入 {name} (学号: {sno}) 的 {cno} 课程新成绩:",
            float(current_grade), 0, 100, 1
        )

        if ok:
            try:
                # 更新数据库
                query = """
                UPDATE Student_Score
                SET Grade = ?
                WHERE Sno = ? AND Cno = ?
                """
                self.cursor.execute(query, (new_grade, sno, cno))
                self.conn.commit()

                # 更新表格
                self.table.setItem(selected_row, 4, QTableWidgetItem(str(new_grade)))

                # 更新状态栏
                self.statusBar().showMessage(f"已更新 {name} 的 {cno} 课程成绩")

            except pyodbc.Error as e:
                logging.error(f"更新成绩错误: {e}")
                QMessageBox.critical(self, "错误", f"更新成绩失败: {e}")

    def delete_student_score(self):
        """删除学生成绩"""
        # 获取选中行
        selected_row = self.table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一条记录")
            return

        # 获取信息
        sno = self.table.item(selected_row, 0).text().strip()
        name = self.table.item(selected_row, 1).text().strip()
        cno = self.table.item(selected_row, 2).text().strip()

        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除 {name} (学号: {sno}) 的 {cno} 课程成绩记录？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # 删除记录
                query = "DELETE FROM Student_Score WHERE Sno = ? AND Cno = ?"
                self.cursor.execute(query, (sno, cno))
                self.conn.commit()

                # 从表格中移除
                self.table.removeRow(selected_row)

                # 更新状态栏
                self.statusBar().showMessage(f"已删除 {name} 的 {cno} 课程成绩记录")

            except pyodbc.Error as e:
                logging.error(f"删除成绩错误: {e}")
                QMessageBox.critical(self, "错误", f"删除成绩失败: {e}")

    def add_course(self):
        """添加课程"""
        # 显示对话框获取课程信息
        course_dialog = QDialog(self)
        course_dialog.setWindowTitle("添加新课程")
        course_dialog.setMinimumWidth(300)
        # 移除问号图标
        course_dialog.setWindowFlags(course_dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.apply_dialog_style(course_dialog)

        layout = QVBoxLayout(course_dialog)
        form_layout = QFormLayout()

        cno_input = QLineEdit()
        cno_input.setPlaceholderText("课程号，如C001")

        # 添加说明标签
        cno_tip = QLabel("课程号必须为一个字母后跟三个数字")
        cno_tip.setStyleSheet("color: #666; font-size: 10px;")

        name_input = QLineEdit()
        name_input.setPlaceholderText("输入课程名称")

        credit_input = QLineEdit()
        credit_input.setPlaceholderText("输入学分数")

        # 添加学分提示
        credit_tip = QLabel("支持小数，如: 2.5, 3.0, 4.5")
        credit_tip.setStyleSheet("color: #666; font-size: 10px;")

        form_layout.addRow("课程号:", cno_input)
        form_layout.addRow("", cno_tip)
        form_layout.addRow("课程名称:", name_input)
        form_layout.addRow("学分:", credit_input)
        form_layout.addRow("", credit_tip)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(course_dialog.accept)
        buttons.rejected.connect(course_dialog.reject)
        layout.addWidget(buttons)
        self.setup_standard_buttons(buttons)

        if course_dialog.exec_() == QDialog.Accepted:
            cno = cno_input.text().strip()
            name = name_input.text().strip()

            try:
                # 验证课程号格式
                is_valid, error_msg = validate_course_number(cno)
                if not is_valid:
                    QMessageBox.warning(self, "格式错误", error_msg)
                    return

                credit = float(credit_input.text().strip())  # 修改为float以支持小数

                if not cno or not name:
                    QMessageBox.warning(self, "输入错误", "课程号和课程名称不能为空")
                    return

                try:
                    # 检查课程号是否已存在
                    self.cursor.execute("SELECT COUNT(*) FROM Course WHERE Cno = ?", (cno,))
                    if self.cursor.fetchone()[0] > 0:
                        QMessageBox.warning(self, "添加失败", "该课程号已存在")
                        return

                    # 添加课程
                    self.cursor.execute("""
                        INSERT INTO Course (Cno, course_name, Credit)
                        VALUES (?, ?, ?)
                    """, (cno, name, credit))

                    self.conn.commit()

                    # 刷新数据
                    self.load_courses()

                    # 更新状态栏
                    self.statusBar().showMessage(f"已添加课程: {name} ({cno})")

                except pyodbc.Error as e:
                    self.conn.rollback()
                    logging.error(f"添加课程错误: {e}")
                    QMessageBox.critical(self, "错误", f"添加课程失败: {e}")

            except ValueError:
                QMessageBox.warning(self, "输入错误", "学分必须是数字")

    def edit_course(self):
        """编辑课程 """
        selected_row = self.course_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一条课程记录")
            return

        # 获取当前课程信息
        old_cno = self.course_table.item(selected_row, 0).text().strip()
        current_name = self.course_table.item(selected_row, 1).text().strip()
        current_credit = self.course_table.item(selected_row, 2).text()

        # 显示编辑对话框
        course_dialog = QDialog(self)
        course_dialog.setWindowTitle(f"编辑课程 - {current_name}")
        course_dialog.setMinimumWidth(300)
        # 移除问号图标
        course_dialog.setWindowFlags(course_dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.apply_dialog_style(course_dialog)

        layout = QVBoxLayout(course_dialog)
        form_layout = QFormLayout()

        # 让课程号可编辑
        cno_input = QLineEdit(old_cno)
        # 添加说明标签
        cno_tip = QLabel("课程号必须为一个字母后跟三个数字（例如：C001）")
        cno_tip.setStyleSheet("color: #666; font-size: 10px;")

        name_input = QLineEdit(current_name)

        credit_input = QLineEdit(current_credit)
        # 添加提示信息
        credit_tip = QLabel("支持小数，如: 2.5, 3.0, 4.5")
        credit_tip.setStyleSheet("color: #666; font-size: 10px;")

        form_layout.addRow("课程号:", cno_input)
        form_layout.addRow("", cno_tip)
        form_layout.addRow("课程名称:", name_input)
        form_layout.addRow("学分:", credit_input)
        form_layout.addRow("", credit_tip)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(course_dialog.accept)
        buttons.rejected.connect(course_dialog.reject)
        layout.addWidget(buttons)
        self.setup_standard_buttons(buttons)

        if course_dialog.exec_() == QDialog.Accepted:
            new_cno = cno_input.text().strip()
            name = name_input.text().strip()
            credit_text = credit_input.text().strip()

            # 验证输入的基本有效性
            if not name:
                QMessageBox.warning(self, "输入错误", "课程名称不能为空")
                return

            # 验证学分值
            try:
                credit = float(credit_text)
                if credit <= 0:
                    QMessageBox.warning(self, "输入错误", "学分必须大于0")
                    return
            except ValueError:
                QMessageBox.warning(self, "输入错误", "学分必须是数字")
                return

            # 根据是否修改课程号分别处理
            is_changing_cno = (new_cno != old_cno)

            # 如果修改了课程号，才进行课程号格式验证
            if is_changing_cno:
                # 验证新课程号格式
                is_valid, error_msg = validate_course_number(new_cno)
                if not is_valid:
                    QMessageBox.warning(self, "课程号格式错误", error_msg)
                    return

                # 检查新课程号是否已被使用
                try:
                    self.cursor.execute("SELECT COUNT(*) FROM Course WHERE Cno = ? AND Cno <> ?",
                                        (new_cno, old_cno))
                    if self.cursor.fetchone()[0] > 0:
                        QMessageBox.warning(self, "编辑失败", "该课程号已被其他课程使用")
                        return
                except pyodbc.Error as e:
                    logging.error(f"查询课程号错误: {e}")
                    QMessageBox.critical(self, "数据库错误", f"检查课程号时出错: {e}")
                    return

            # 执行数据库更新操作
            try:
                if is_changing_cno:
                    # 需要处理外键关系的课程号修改
                    self.cursor.execute("BEGIN TRANSACTION")

                    # 1. 先添加新课程记录
                    self.cursor.execute("""
                        INSERT INTO Course (Cno, course_name, Credit)
                        VALUES (?, ?, ?)
                    """, (new_cno, name, credit))

                    # 2. 更新成绩记录引用新课程号
                    self.cursor.execute("""
                        UPDATE Student_Score 
                        SET Cno = ?
                        WHERE Cno = ?
                    """, (new_cno, old_cno))

                    # 3. 删除旧课程记录
                    self.cursor.execute("DELETE FROM Course WHERE Cno = ?", (old_cno,))

                    self.conn.commit()

                    # 提示信息
                    change_type = "课程号、课程名称和学分"
                else:
                    # 仅更新名称和学分，不涉及课程号变更
                    self.cursor.execute("""
                        UPDATE Course 
                        SET course_name = ?, Credit = ? 
                        WHERE Cno = ?
                    """, (name, credit, old_cno))
                    self.conn.commit()

                    # 提示信息
                    if name != current_name and credit != float(current_credit):
                        change_type = "课程名称和学分"
                    elif name != current_name:
                        change_type = "课程名称"
                    else:
                        change_type = "学分"

                # 刷新数据
                self.load_courses()
                # 如果更改了课程号，也要刷新成绩表
                if is_changing_cno:
                    self.load_students()

                # 格式化学分显示
                formatted_credit = f"{credit:.1f}" if credit != int(credit) else str(int(credit))

                # 显示成功消息
                self.statusBar().showMessage(f"已更新课程 {change_type}: {name} ({new_cno}), 学分: {formatted_credit}")

            except pyodbc.Error as e:
                # 发生错误时回滚事务
                if is_changing_cno:
                    self.conn.rollback()
                logging.error(f"更新课程错误: {e}")
                QMessageBox.critical(self, "数据库错误", f"更新课程失败: {e}")

    def delete_course(self):
        """删除课程"""
        # 获取选中行
        selected_row = self.course_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一条课程记录")
            return

        # 获取课程信息
        cno = self.course_table.item(selected_row, 0).text().strip()
        name = self.course_table.item(selected_row, 1).text().strip()

        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除课程 {name} ({cno})？\n\n注意：删除课程将同时删除所有学生的该课程成绩记录！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # 删除课程前先删除相关成绩记录 (外键约束)
                self.cursor.execute("DELETE FROM Student_Score WHERE Cno = ?", (cno,))

                # 删除课程
                self.cursor.execute("DELETE FROM Course WHERE Cno = ?", (cno,))

                self.conn.commit()

                # 刷新数据
                self.course_table.removeRow(selected_row)
                self.load_students()  # 也要刷新学生成绩表

                # 更新状态栏
                self.statusBar().showMessage(f"已删除课程: {name} ({cno})")

            except pyodbc.Error as e:
                self.conn.rollback()
                logging.error(f"删除课程错误: {e}")
                QMessageBox.critical(self, "错误", f"删除课程失败: {e}")

    def import_excel(self):
        """从Excel导入数据 """
        # 打开文件对话框选择Excel文件
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择Excel文件", "", "Excel文件 (*.xlsx *.xls)"
        )

        if not file_path:
            return

        try:
            # 创建进度对话框
            progress = QDialog(self)
            progress.setWindowTitle("导入中...")
            progress.setFixedSize(300, 100)
            self.apply_dialog_style(progress)

            progress_layout = QVBoxLayout(progress)
            progress_label = QLabel("正在读取Excel文件，请稍候...")
            progress_layout.addWidget(progress_label)
            progress.setModal(True)
            progress.show()
            QApplication.processEvents()  # 更新UI

            # 读取Excel文件
            df = pd.read_excel(file_path)

            # 更新进度提示
            progress_label.setText("正在分析数据...")
            QApplication.processEvents()

            # 获取当前活动的选项卡
            tab_index = self.findChild(QTabWidget).currentIndex()

            # 根据不同选项卡处理不同类型的数据
            if tab_index == 0:  # 学生成绩选项卡
                # 检查必要的列
                required_columns = ["学号", "姓名", "课程号", "课程名称", "学分", "成绩"]

                # 检查表头，支持多种表头格式
                header_mappings = {
                    "学号": ["学号", "学生学号", "学生号", "sno", "student_id", "学生id"],
                    "姓名": ["姓名", "学生姓名", "name", "student_name"],
                    "课程号": ["课程号", "课号", "课程编号", "cno", "course_id"],
                    "课程名称": ["课程名称", "课程名", "课程", "course_name", "course"],
                    "学分": ["学分", "分值", "学分值", "credit"],
                    "成绩": ["成绩", "得分", "分数", "grade", "score"]
                }

                # 尝试匹配列名
                matched_columns = {}
                for required_col, possible_names in header_mappings.items():
                    for col in df.columns:
                        if str(col).lower() in [name.lower() for name in possible_names]:
                            matched_columns[required_col] = col
                            break

                missing_columns = [col for col in required_columns if col not in matched_columns]

                if missing_columns:
                    progress.close()
                    QMessageBox.warning(
                        self,
                        "格式错误",
                        f"Excel文件缺少必要的列: {', '.join(missing_columns)}\n"
                        f"请确保Excel包含以下列: {', '.join(required_columns)}"
                    )
                    return

                # 创建一个新的DataFrame使用标准列名
                standardized_df = pd.DataFrame()
                for std_col, excel_col in matched_columns.items():
                    standardized_df[std_col] = df[excel_col]

                # 预处理数据
                standardized_df = standardized_df.fillna("")  # 将NaN替换为空字符串

                # 数据验证
                errors = []
                for index, row in standardized_df.iterrows():
                    # 验证学号
                    if not row["学号"] or str(row["学号"]).strip() == "":
                        errors.append(f"第 {index + 2} 行: 学号不能为空")

                    # 验证姓名
                    if not row["姓名"] or str(row["姓名"]).strip() == "":
                        errors.append(f"第 {index + 2} 行: 姓名不能为空")

                    # 验证课程号
                    if not row["课程号"] or str(row["课程号"]).strip() == "":
                        errors.append(f"第 {index + 2} 行: 课程号不能为空")

                    # 验证学分
                    try:
                        if row["学分"] != "":
                            credit = float(row["学分"])
                            if credit <= 0:
                                errors.append(f"第 {index + 2} 行: 学分必须大于0")
                    except:
                        errors.append(f"第 {index + 2} 行: 学分必须是数字")

                    # 验证成绩
                    try:
                        if row["成绩"] != "":
                            grade = float(row["成绩"])
                            if grade < 0 or grade > 100:
                                errors.append(f"第 {index + 2} 行: 成绩必须在0-100之间")
                    except:
                        errors.append(f"第 {index + 2} 行: 成绩必须是数字")

                if errors:
                    progress.close()
                    error_msg = "\n".join(errors[:10])
                    if len(errors) > 10:
                        error_msg += f"\n... 等共 {len(errors)} 个错误"
                    QMessageBox.warning(self, "数据验证错误", f"Excel数据存在以下问题:\n{error_msg}")
                    return

                # 关闭进度对话框
                progress.close()

                # 导入确认对话框
                confirm_dialog = QDialog(self)
                confirm_dialog.setWindowTitle("确认导入")
                confirm_dialog.setMinimumSize(600, 400)

                self.apply_dialog_style(confirm_dialog)
                layout = QVBoxLayout(confirm_dialog)

                info_label = QLabel(f"即将导入 {len(standardized_df)} 条学生成绩记录，请选择导入选项：")
                layout.addWidget(info_label)

                # 导入选项
                options_group = QGroupBox("导入选项")
                options_layout = QVBoxLayout()

                self.update_existing = QCheckBox("更新已存在的成绩记录")
                self.update_existing.setChecked(True)

                self.add_missing_students = QCheckBox("自动添加不存在的学生")
                self.add_missing_students.setChecked(True)

                self.add_missing_courses = QCheckBox("自动添加不存在的课程")
                self.add_missing_courses.setChecked(True)

                options_layout.addWidget(self.update_existing)
                options_layout.addWidget(self.add_missing_students)
                options_layout.addWidget(self.add_missing_courses)

                options_group.setLayout(options_layout)
                layout.addWidget(options_group)

                # 预览表格
                preview_label = QLabel("数据预览：")
                layout.addWidget(preview_label)

                preview_table = QTableWidget()
                preview_table.setColumnCount(len(standardized_df.columns))
                preview_table.setHorizontalHeaderLabels(standardized_df.columns.tolist())

                # 仅显示前10行数据
                max_rows = min(10, len(standardized_df))
                preview_table.setRowCount(max_rows)

                for i in range(max_rows):
                    for j, col in enumerate(standardized_df.columns):
                        value = standardized_df.iloc[i, j]
                        if pd.isna(value):
                            value = ""
                        item = QTableWidgetItem(str(value))
                        preview_table.setItem(i, j, item)

                preview_table.resizeColumnsToContents()
                layout.addWidget(preview_table)

                # 按钮
                buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttons.accepted.connect(confirm_dialog.accept)
                buttons.rejected.connect(confirm_dialog.reject)
                layout.addWidget(buttons)
                self.setup_standard_buttons(buttons)

                if confirm_dialog.exec_() == QDialog.Accepted:
                    # 重新打开进度对话框
                    progress = QDialog(self)
                    progress.setWindowTitle("导入中...")
                    progress.setFixedSize(300, 100)
                    progress_layout = QVBoxLayout(progress)
                    progress_label = QLabel("正在导入数据，请稍候...")
                    progress_layout.addWidget(progress_label)
                    progress.setModal(True)
                    progress.show()
                    QApplication.processEvents()  # 更新UI

                    # 开始导入
                    imported = 0
                    updated = 0
                    errors = 0

                    try:
                        # 开始导入事务
                        self.cursor.execute("BEGIN TRANSACTION")

                        for i, row in standardized_df.iterrows():
                            # 更新进度显示
                            if i % 10 == 0:
                                progress_label.setText(f"正在处理第 {i + 1}/{len(standardized_df)} 条记录...")
                                QApplication.processEvents()

                            try:
                                # 清理数据并转换类型
                                sno = str(row["学号"]).strip()
                                name = str(row["姓名"]).strip()
                                cno = str(row["课程号"]).strip()

                                is_valid, error_msg = validate_course_number(cno)
                                if not is_valid:
                                    logging.warning(f"行 {i + 2}：课程号 '{cno}' 格式错误 - {error_msg}")
                                    errors += 1
                                    continue

                                # 处理可能的空值或非数字值
                                try:
                                    credit = float(row["学分"]) if row["学分"] != "" else 0.0
                                except:
                                    credit = 0.0

                                try:
                                    grade = float(row["成绩"]) if row["成绩"] != "" else 0.0
                                except:
                                    grade = 0.0

                                # 检查学生是否存在
                                self.cursor.execute("SELECT COUNT(*) FROM Student WHERE Sno = ?", (sno,))
                                if self.cursor.fetchone()[0] == 0:
                                    if self.add_missing_students.isChecked():
                                        # 添加学生
                                        self.cursor.execute(
                                            "INSERT INTO Student (Sno, name) VALUES (?, ?)",
                                            (sno, name)
                                        )
                                    else:
                                        logging.warning(f"学生 {sno} 不存在且未选择自动添加")
                                        errors += 1
                                        continue

                                # 检查课程是否存在
                                self.cursor.execute("SELECT COUNT(*) FROM Course WHERE Cno = ?", (cno,))
                                if self.cursor.fetchone()[0] == 0:
                                    if self.add_missing_courses.isChecked():
                                        # 添加课程
                                        course_name = str(row["课程名称"]).strip() if "课程名称" in row and row[
                                            "课程名称"] != "" else f"[导入]课程{cno}"
                                        self.cursor.execute(
                                            "INSERT INTO Course (Cno, course_name, Credit) VALUES (?, ?, ?)",
                                            (cno, course_name, credit)
                                        )
                                    else:
                                        logging.warning(f"课程 {cno} 不存在且未选择自动添加")
                                        errors += 1
                                        continue

                                # 检查成绩记录是否存在
                                self.cursor.execute(
                                    "SELECT COUNT(*) FROM Student_Score WHERE Sno = ? AND Cno = ?",
                                    (sno, cno)
                                )

                                if self.cursor.fetchone()[0] > 0:
                                    if self.update_existing.isChecked():
                                        # 更新成绩
                                        self.cursor.execute(
                                            "UPDATE Student_Score SET Grade = ? WHERE Sno = ? AND Cno = ?",
                                            (grade, sno, cno)
                                        )
                                        updated += 1
                                else:
                                    # 添加成绩记录
                                    self.cursor.execute(
                                        "INSERT INTO Student_Score (Sno, Cno, Grade) VALUES (?, ?, ?)",
                                        (sno, cno, grade)
                                    )
                                    imported += 1

                            except Exception as e:
                                logging.error(f"导入行数据错误 (行 {i + 2}): {e}")
                                errors += 1

                        # 提交事务
                        self.cursor.execute("COMMIT")
                        progress.close()

                        # 刷新数据
                        self.load_students()
                        self.load_courses()

                        # 显示导入结果
                        QMessageBox.information(
                            self,
                            "导入完成",
                            f"导入成功！\n\n"
                            f"新增记录: {imported}\n"
                            f"更新记录: {updated}\n"
                            f"错误记录: {errors}"
                        )

                        self.statusBar().showMessage(f"成功导入 {imported} 条记录，更新 {updated} 条记录")

                    except Exception as e:
                        # 回滚事务
                        self.cursor.execute("ROLLBACK")
                        progress.close()
                        logging.error(f"导入数据错误: {e}")
                        QMessageBox.critical(self, "导入错误", f"导入数据失败: {e}")
                        self.statusBar().showMessage("导入失败")

            else:
                # 其他选项卡的导入功能
                progress.close()
                QMessageBox.information(
                    self,
                    "功能提示",
                    "请切换到\"学生成绩\"选项卡使用导入功能"
                )

        except Exception as e:
            # 确保进度对话框关闭
            try:
                progress.close()
            except:
                pass
            logging.error(f"读取Excel文件错误: {e}")
            QMessageBox.critical(self, "文件错误", f"无法读取Excel文件: {e}")

    def export_excel(self):
        """导出数据到Excel """
        # 打开文件对话框选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存Excel文件", "", "Excel文件 (*.xlsx)"
        )

        if not file_path:
            return

        # 如果文件路径没有.xlsx后缀，添加它
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'

        try:
            # 创建一个等待对话框
            progress = QDialog(self)
            progress.setWindowTitle("导出中...")
            progress.setFixedSize(300, 100)
            progress_layout = QVBoxLayout(progress)
            progress_label = QLabel("正在导出数据，请稍候...")
            progress_layout.addWidget(progress_label)
            progress.setModal(True)
            progress.show()
            QApplication.processEvents()  # 更新UI

            # 获取数据 - 根据当前视图决定导出内容
            tab_index = self.findChild(QTabWidget).currentIndex()

            if tab_index == 0:  # 学生成绩
                # 创建要导出的学生成绩数据
                data = []
                headers = ["学号", "姓名", "课程号", "课程名称", "学分", "成绩"]

                # 执行查询获取完整关联数据
                self.cursor.execute("""
                    SELECT s.Sno, s.name, sc.Cno, c.course_name, c.Credit, sc.Grade
                    FROM Student_Score sc
                    JOIN Student s ON sc.Sno = s.Sno
                    JOIN Course c ON sc.Cno = c.Cno
                    ORDER BY s.Sno, sc.Cno
                """)

                records = self.cursor.fetchall()
                for record in records:
                    row_data = []
                    for value in record:
                        # 处理小数格式
                        if isinstance(value, float) and value == int(value):
                            value = int(value)
                        row_data.append(value)
                    data.append(row_data)

            elif tab_index == 1:  # 学生管理
                # 导出学生信息
                headers = ["学号", "姓名", "已修课程数", "平均成绩"]
                data = []

                # 执行查询获取学生数据及统计信息
                self.cursor.execute("""
                    SELECT s.Sno, s.name, 
                           COUNT(sc.Cno) as course_count,
                           AVG(sc.Grade) as avg_grade
                    FROM Student s
                    LEFT JOIN Student_Score sc ON s.Sno = sc.Sno
                    GROUP BY s.Sno, s.name
                    ORDER BY s.Sno
                """)

                records = self.cursor.fetchall()
                for record in records:
                    row_data = []
                    for i, value in enumerate(record):
                        # 处理平均成绩格式
                        if i == 3 and value is not None:  # 平均成绩
                            value = round(value, 2)
                        elif value is None:
                            value = 0
                        row_data.append(value)
                    data.append(row_data)

            else:  # 课程管理
                # 导出课程信息
                headers = ["课程号", "课程名称", "学分", "选课人数", "平均成绩"]
                data = []

                # 执行查询获取课程数据及统计信息
                self.cursor.execute("""
                    SELECT c.Cno, c.course_name, c.Credit,
                           COUNT(sc.Sno) as student_count,
                           AVG(sc.Grade) as avg_grade
                    FROM Course c
                    LEFT JOIN Student_Score sc ON c.Cno = sc.Cno
                    GROUP BY c.Cno, c.course_name, c.Credit
                    ORDER BY c.Cno
                """)

                records = self.cursor.fetchall()
                for record in records:
                    row_data = []
                    for i, value in enumerate(record):
                        # 处理小数格式
                        if i == 2 and value is not None:  # 学分
                            if value == int(value):
                                value = int(value)
                        elif i == 4 and value is not None:  # 平均成绩
                            value = round(value, 2)
                        elif value is None:
                            value = 0
                        row_data.append(value)
                    data.append(row_data)

            # 创建DataFrame
            df = pd.DataFrame(data, columns=headers)

            # 使用ExcelWriter增强Excel格式
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='数据')

                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['数据']

                # 创建表头样式
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                centered = Alignment(horizontal="center", vertical="center")
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 应用样式到表头
                for col_num, column in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = centered
                    cell.border = border

                # 设置列宽
                for i, column in enumerate(df.columns):
                    column_width = max(len(str(column)), df[column].astype(str).map(len).max())
                    worksheet.column_dimensions[chr(65 + i)].width = column_width + 4

                # 应用边框和居中对齐到所有数据单元格
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")

            # 关闭进度对话框
            progress.close()

            # 成功提示
            QMessageBox.information(self, "导出成功", f"成功将 {len(data)} 条记录导出到:\n{file_path}")
            self.statusBar().showMessage(f"成功导出 {len(data)} 条记录到 {file_path}")

        except Exception as e:
            logging.error(f"导出Excel错误: {e}")
            QMessageBox.critical(self, "导出错误", f"导出数据失败: {e}")

    def show_about(self):
        """显示关于对话框"""
        QMessageBox.about(
            self,
            "关于学生信息管理系统",
            "学生信息管理系统 v1.0\n"
            "基于PyQt5和SQL Server开发\n"
            "作者：甄城 U202316703\n"
            "支持学生、课程和成绩管理"
        )

    def closeEvent(self, event):
        """关闭窗口时关闭数据库连接"""
        if hasattr(self, 'conn') and self.conn:
            self.conn.close()
        event.accept()

    def load_student_list(self):
        """加载学生列表"""
        try:
            # 执行查询
            query = """
            SELECT s.Sno, s.name, COUNT(sc.Cno) as course_count
            FROM Student s
            LEFT JOIN Student_Score sc ON s.Sno = sc.Sno
            GROUP BY s.Sno, s.name
            ORDER BY s.Sno
            """
            self.cursor.execute(query)
            students = self.cursor.fetchall()

            # 清空表格
            self.student_table.setRowCount(0)

            # 添加数据到表格
            for i, student in enumerate(students):
                self.student_table.insertRow(i)
                for j, value in enumerate(student):
                    # 修复空格问题
                    if j in [0, 1] and value is not None:
                        value = str(value).strip()
                    self.student_table.setItem(i, j, QTableWidgetItem(str(value)))

            # 更新状态栏
            self.statusBar().showMessage(f"已加载 {len(students)} 名学生")

        except pyodbc.Error as e:
            logging.error(f"加载学生列表错误: {e}")
            QMessageBox.warning(self, "错误", f"加载学生列表失败: {e}")

    def search_student(self):
        """搜索学生"""
        search_text = self.student_search_input.text().strip()

        if not search_text:
            self.load_student_list()
            return

        try:
            # 执行查询
            query = """
            SELECT s.Sno, s.name, COUNT(sc.Cno) as course_count
            FROM Student s
            LEFT JOIN Student_Score sc ON s.Sno = sc.Sno
            WHERE s.Sno LIKE ? OR s.name LIKE ?
            GROUP BY s.Sno, s.name
            ORDER BY s.Sno
            """
            self.cursor.execute(query, (f"%{search_text}%", f"%{search_text}%"))
            students = self.cursor.fetchall()

            # 清空表格
            self.student_table.setRowCount(0)

            # 添加数据到表格
            for i, student in enumerate(students):
                self.student_table.insertRow(i)
                for j, value in enumerate(student):
                    # 修复空格问题
                    if j in [0, 1] and value is not None:
                        value = str(value).strip()
                    self.student_table.setItem(i, j, QTableWidgetItem(str(value)))

            # 更新状态栏
            self.statusBar().showMessage(f"找到 {len(students)} 名匹配的学生")

        except pyodbc.Error as e:
            logging.error(f"搜索学生错误: {e}")
            QMessageBox.warning(self, "错误", f"搜索学生失败: {e}")

    def reset_student_search(self):
        """重置学生搜索"""
        self.student_search_input.clear()
        self.load_student_list()

    def search_course(self):
        """搜索课程"""
        search_text = self.course_search_input.text().strip()

        if not search_text:
            self.load_courses()
            return

        try:
            # 执行查询
            query = """
            SELECT Cno, course_name, Credit
            FROM Course
            WHERE Cno LIKE ? OR course_name LIKE ?
            ORDER BY Cno
            """
            self.cursor.execute(query, (f"%{search_text}%", f"%{search_text}%"))
            courses = self.cursor.fetchall()

            # 清空表格
            self.course_table.setRowCount(0)

            # 添加数据到表格
            for i, course in enumerate(courses):
                self.course_table.insertRow(i)
                for j, value in enumerate(course):
                    # 修复空格问题
                    if j in [0, 1] and value is not None:
                        value = str(value).strip()
                    self.course_table.setItem(i, j, QTableWidgetItem(str(value)))

            # 更新状态栏
            self.statusBar().showMessage(f"找到 {len(courses)} 门匹配的课程")

        except pyodbc.Error as e:
            logging.error(f"搜索课程错误: {e}")
            QMessageBox.warning(self, "错误", f"搜索课程失败: {e}")

    def reset_course_search(self):
        """重置课程搜索"""
        self.course_search_input.clear()
        self.load_courses()

    def show_student_context_menu(self, position):
        """显示学生右键菜单"""
        if not self.student_table.selectedItems():
            return

        context_menu = QMenu(self)

        edit_action = context_menu.addAction("编辑学生")
        view_scores_action = context_menu.addAction("查看成绩")
        delete_action = context_menu.addAction("删除学生")

        edit_action.triggered.connect(self.edit_student)
        view_scores_action.triggered.connect(self.view_student_scores)
        delete_action.triggered.connect(self.delete_student)

        context_menu.exec_(self.student_table.mapToGlobal(position))

    def add_student(self):
        """添加学生"""
        dialog = QDialog(self)
        dialog.setWindowTitle("添加学生")
        dialog.setMinimumWidth(300)

        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.apply_dialog_style(dialog)

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()

        sno_input = QLineEdit()
        sno_input.setPlaceholderText("输入学号")

        name_input = QLineEdit()
        name_input.setPlaceholderText("输入姓名")

        form_layout.addRow("学号:", sno_input)
        form_layout.addRow("姓名:", name_input)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        self.setup_standard_buttons(buttons)

        if dialog.exec_() == QDialog.Accepted:
            sno = sno_input.text().strip()
            name = name_input.text().strip()

            if not sno or not name:
                QMessageBox.warning(self, "输入错误", "学号和姓名不能为空")
                return

            try:
                # 检查学号是否已存在
                self.cursor.execute("SELECT COUNT(*) FROM Student WHERE Sno = ?", (sno,))
                if self.cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "添加失败", "该学号已存在")
                    return

                # 添加学生
                self.cursor.execute("INSERT INTO Student (Sno, name) VALUES (?, ?)", (sno, name))
                self.conn.commit()

                # 刷新数据
                self.load_student_list()

                # 更新状态栏
                self.statusBar().showMessage(f"已添加学生: {name} ({sno})")

            except pyodbc.Error as e:
                self.conn.rollback()
                logging.error(f"添加学生错误: {e}")
                QMessageBox.critical(self, "错误", f"添加学生失败: {e}")

    def edit_student(self):
        """编辑学生"""
        selected_row = self.student_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一名学生")
            return

        # 获取当前学生信息
        old_sno = self.student_table.item(selected_row, 0).text().strip()
        current_name = self.student_table.item(selected_row, 1).text().strip()

        dialog = QDialog(self)
        dialog.setWindowTitle(f"编辑学生")
        dialog.setMinimumWidth(300)
        # 移除问号图标
        dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
        self.apply_dialog_style(dialog)  # 使用新添加的样式方法

        layout = QVBoxLayout(dialog)
        form_layout = QFormLayout()

        # 将学号从标签改为输入框
        sno_input = QLineEdit(old_sno)
        name_input = QLineEdit(current_name)

        form_layout.addRow("学号:", sno_input)
        form_layout.addRow("姓名:", name_input)

        layout.addLayout(form_layout)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec_() == QDialog.Accepted:
            new_sno = sno_input.text().strip()
            name = name_input.text().strip()

            if not new_sno or not name:
                QMessageBox.warning(self, "输入错误", "学号和姓名不能为空")
                return

            try:
                if new_sno != old_sno:
                    # 检查新学号是否已存在
                    self.cursor.execute("SELECT COUNT(*) FROM Student WHERE Sno = ? AND Sno <> ?",
                                        (new_sno, old_sno))
                    if self.cursor.fetchone()[0] > 0:
                        QMessageBox.warning(self, "编辑失败", "该学号已被其他学生使用")
                        return

                    # 开始事务
                    self.cursor.execute("BEGIN TRANSACTION")

                    # 更新成绩记录中的学号
                    self.cursor.execute("UPDATE Student_Score SET Sno = ? WHERE Sno = ?",
                                        (new_sno, old_sno))

                    # 更新学生信息
                    self.cursor.execute("UPDATE Student SET Sno = ?, name = ? WHERE Sno = ?",
                                        (new_sno, name, old_sno))
                else:
                    # 只更新姓名
                    self.cursor.execute("UPDATE Student SET name = ? WHERE Sno = ?",
                                        (name, old_sno))

                self.conn.commit()

                # 刷新数据
                self.load_student_list()
                self.load_students()  # 刷新成绩表中的学生信息

                # 更新状态栏
                self.statusBar().showMessage(f"已更新学生信息，学号: {new_sno}, 姓名: {name}")

            except pyodbc.Error as e:
                # 发生错误时回滚事务
                self.conn.rollback()
                logging.error(f"更新学生错误: {e}")
                QMessageBox.critical(self, "错误", f"更新学生失败: {e}")

    def delete_student(self):
        """删除学生"""
        selected_row = self.student_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一名学生")
            return

        sno = self.student_table.item(selected_row, 0).text().strip()
        name = self.student_table.item(selected_row, 1).text().strip()

        # 确认删除
        reply = QMessageBox.question(
            self,
            "确认删除",
            f"确定要删除学生 {name} ({sno})？\n\n注意：删除学生将同时删除该学生的所有成绩记录！",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # 开始事务
                self.cursor.execute("BEGIN TRANSACTION")

                # 先删除成绩记录
                self.cursor.execute("DELETE FROM Student_Score WHERE Sno = ?", (sno,))

                # 删除学生
                self.cursor.execute("DELETE FROM Student WHERE Sno = ?", (sno,))

                # 提交事务
                self.cursor.execute("COMMIT")

                # 刷新数据
                self.student_table.removeRow(selected_row)
                self.load_students()  # 也要刷新成绩表

                # 更新状态栏
                self.statusBar().showMessage(f"已删除学生: {name} ({sno})")

            except pyodbc.Error as e:
                # 回滚事务
                self.cursor.execute("ROLLBACK")
                logging.error(f"删除学生错误: {e}")
                QMessageBox.critical(self, "错误", f"删除学生失败: {e}")

    def view_student_scores(self):
        """查看学生成绩"""
        selected_row = self.student_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "警告", "请先选择一名学生")
            return


        sno = self.student_table.item(selected_row, 0).text().strip()
        name = self.student_table.item(selected_row, 1).text().strip()

        try:
            # 查询学生成绩
            query = """
            SELECT sc.Cno, c.course_name, c.Credit, sc.Grade
            FROM Student_Score sc
            JOIN Course c ON sc.Cno = c.Cno
            WHERE sc.Sno = ?
            ORDER BY sc.Cno
            """
            self.cursor.execute(query, (sno,))
            scores = self.cursor.fetchall()

            if not scores:
                QMessageBox.information(self, "提示", f"学生 {name} ({sno}) 暂无成绩记录")
                return

            # 创建成绩对话框
            dialog = QDialog(self)
            dialog.setWindowTitle(f"{name} ({sno}) 的成绩记录")
            dialog.setMinimumSize(500, 400)

            dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
            self.apply_dialog_style(dialog)
            layout = QVBoxLayout(dialog)

            # 成绩表格
            score_table = QTableWidget(len(scores), 4)
            score_table.setHorizontalHeaderLabels(["课程号", "课程名称", "学分", "成绩"])
            score_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            score_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

            # 添加数据
            total_credit = 0
            weighted_sum = 0

            for i, score in enumerate(scores):
                for j, value in enumerate(score):
                    # 修复空格问题
                    if j in [0, 1] and value is not None:
                        value = str(value).strip()
                    score_table.setItem(i, j, QTableWidgetItem(str(value)))

                # 计算加权分数
                credit = float(score[2])
                grade = float(score[3])
                total_credit += credit
                weighted_sum += credit * grade

            layout.addWidget(score_table)

            # 统计信息
            stats_layout = QFormLayout()

            stats_layout.addRow("课程总数:", QLabel(str(len(scores))))
            stats_layout.addRow("总学分:", QLabel(str(total_credit)))

            # 计算平均成绩（加权）
            if total_credit > 0:
                avg_grade = weighted_sum / total_credit
                avg_label = QLabel(f"{avg_grade:.2f}")
                # 设置颜色
                if avg_grade >= 90:
                    avg_label.setStyleSheet("color: #4caf50; font-weight: bold;")  # 绿色
                elif avg_grade >= 80:
                    avg_label.setStyleSheet("color: #2196f3; font-weight: bold;")  # 蓝色
                elif avg_grade >= 60:
                    avg_label.setStyleSheet("color: #ff9800; font-weight: bold;")  # 橙色
                else:
                    avg_label.setStyleSheet("color: #f44336; font-weight: bold;")  # 红色
            else:
                avg_label = QLabel("N/A")

            stats_layout.addRow("平均成绩(加权):", avg_label)

            layout.addLayout(stats_layout)

            # 关闭按钮
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.accept)
            layout.addWidget(close_btn)

            dialog.exec_()

        except pyodbc.Error as e:
            logging.error(f"查询学生成绩错误: {e}")
            QMessageBox.critical(self, "错误", f"查询学生成绩失败: {e}")


if __name__ == "__main__":
    # 设置异常处理
    def exception_hook(exctype, value, traceback):
        print(exctype, value, traceback)
        sys.__excepthook__(exctype, value, traceback)
        sys.exit(1)


    sys.excepthook = exception_hook

    app = QApplication(sys.argv)

    app.setStyleSheet("""
            QDialogButtonBox QPushButton {
                color: black !important;
                background-color: #e0e0e0;
                font-weight: bold;
                min-width: 80px;
                padding: 5px 15px;
                border: 2px solid #9e9e9e;
                border-radius: 4px;
            }
            QDialogButtonBox QPushButton:hover {
                background-color: #d0d0d0;
            }
            QDialogButtonBox QPushButton[text="OK"], 
            QDialogButtonBox QPushButton[text="确定"],
            QDialogButtonBox QPushButton[text="Yes"],
            QDialogButtonBox QPushButton[text="是"] {
                color: white !important;
                background-color: #4caf50;
                border: 2px solid #388e3c;
            }
            QMessageBox QPushButton {
                color: black !important;
                background-color: #e0e0e0;
                min-width: 80px;
                padding: 5px 15px;
            }
        """)

    # 设置应用样式
    app.setStyle("Fusion")

    # 设置应用图标
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "image", "pic.ico")
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))

    window = StudentManagementSystem()
    window.show()

    sys.exit(app.exec_())